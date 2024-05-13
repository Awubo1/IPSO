import numpy as np
import matplotlib.pyplot as plt
import xlrd
import random
import math
from openpyxl import Workbook
from scipy.special import gamma
'''

'''
#########
class   GA_PSO():#18 5
    def __init__(self,length, dim,pop, max_iter):  #   长度  维度，群体数量，迭代次数
        self.dim = dim  # 基因维度#
        self.length = length# #基因长度
        self.pop = pop  # 粒子数量# 这里是初始化粒子群，20个
        self.max_iter = max_iter  # 迭代次数
        self.population = []  # 父代种群
        self.V=[]
        self.new_popu = []  # 选择算子操作过后的新种群
        self.new_V=[]
        self.fitness=[]    ##种群中所有基因的适应度函数
        self.alpha=0.5#负载权重因子
        self.elite=[]    ##精英基因
        self.elite_index=0   ##精英基因的序号
        self.iter_bfit=[]     ##每次迭代中的最小的fitness
        self.pbest=[]
        self.wirespeed=3*10^5#电磁波传输速率Km/s
        # self.w = 1  # 设置惯性权重
        self.cmax=2.1#学习系数最大值
        self.cmin = 0.9# 学习系数最小值
        # self.c1 = 2  # 设置个体学习系数
        # self.c2 = 2  # 设置全局学习系数
        self.r1 = None
        self.r2 = None
        self.max_val = int(4)  # 限定最大速度为0.5
       # t_server = offloading_ratio * task_size / (self.f_mec / self.s)
        # self.best_fitness = float(9e10)  # 初始化适应度的值
        self.w_ini=0.9
        self.w_end=0.1
        self.now_iter=0
        ####
        self.mec=9e10 #mec计算频率90GHz
        self.s=random.randint(20, 40)#单位bit处理所需cpu圈数
        self.s1=random.randint(1.5e8, 4e8)#给每个任务提供的cpu圈数
        self.edge_num = 0
        # self.C = 300000  # 20Mhz=
        self.several_groups = 0  # 初始化随机
        self.tran_v=10*1024  ##MB/s=1024kBs
        self.ing_group = 0  # 第几个任务总数序号
        # self.total_num = 0
        self.bs_local = []
        self.edge_local = []
        self.work_size_total = []
        # self.work_num_total = []
        self.edge_index = []
        self.matrix = []
        self.path = []
        self.dis = []
        self.edge_task = []
        self.tran_time = []
        self.edge_use = []
        ####
        self.total_L = []
        self.total_t = []


    def read(self):
        #读取坐标数据
        # xlsx1= xlrd.open_workbook(r"/home/u/Desktop/part2ing_algor/bs_end.xlsx")
        xlsx1 = xlrd.open_workbook(r"D:\边缘计算\data_and_code\code\jizhan.xlsx")
        bs_= xlsx1.sheets()[0]
        # edge_ = xlsx1.sheets()[1]

        self.bs_num= bs_.nrows-1#行数
        b=bs_.ncols-1   ##列数

        # self.bs_local=[]
        for i in range(self.bs_num):
            local=[]
            local = bs_.row_values(i+1)
            # print(sheet1_nrow)
            self.bs_local.append((local[b-1],local[b]))
        # print(self.bs_local)
        ####
        # xlsx2 = xlrd.open_workbook(r"/home/u/Desktop/part2ing_algor/edge_location.xlsx")
        xlsx2 = xlrd.open_workbook(r"D:\边缘计算\data_and_code\code\fuwuqi.xlsx")
        edge_ = xlsx2.sheets()[0]

        self.edge_num = edge_.nrows

        local=[]
        for i in range(self.edge_num):
            local= edge_.row_values(i)
            # print(sheet1_nrow)
            self.edge_local.append((local[0],local[1]))
        # print(self.edge_local)


        for i in range(len(self.edge_local)):
            for j in range(len(self.bs_local)):
                if self.edge_local[i] == self.bs_local[j]:
                    self.edge_index.append(j+1)
        self.edge_index = np.array(self.edge_index)  #edge_index排序
        # # print(edge_index)
        self.edge_index = list(np.sort(self.edge_index))
        # print('self.edge_index:',self.edge_index)  ##服务器编号等于数组序号+1

        ##
        # xlsx3 = xlrd.open_workbook(r"/home/u/Desktop/part2ing_algor/task.xlsx")
        xlsx3 = xlrd.open_workbook(r"D:\边缘计算\data_and_code\code\1500task200-1000.xlsx")
        date_size = xlsx3.sheets()[0]
        # date_size = xlsx1.sheets()[2]
        size_ncol=date_size.ncols
        self.several_groups=int(size_ncol/self.bs_num)
        #所有基站任务总数有几个，分为了几个组
        m = 0
        for i in range(self.several_groups):
            groups = []
            # groups_num = []
            for j in range(self.bs_num):
                a = date_size.col_values(m)
                del a[0]
                a = list(filter(lambda x: x != '', a))
                # print(len(a))
                m += 1
                groups.append(a)
                # print(groups)
            self.work_size_total.append(groups)

        ##matrix
        # xlsx4 = xlrd.open_workbook(r"/home/u/Desktop/part2ing_algor/graph_matrix.xlsx")
        xlsx4 = xlrd.open_workbook(r"D:\边缘计算\data_and_code\code\e.xlsx")
        graph_matrix = xlsx4.sheets()[0]
        matrix_ncol = graph_matrix.ncols
        # matrix_nrow = matrix.nrows
        # print(matrix_ncol)
        # print(matrix_nrow)
        for i in range(matrix_ncol):
            a = graph_matrix.row_values(i)
            self.matrix.append(a)
        # print('self.matrix:')
        # print(self.matrix)

    #s:出发点  出发点，终点为基站服务器序号(不需要减一)
    def Dijkstra(self,network, s, d):  # 迪杰斯特拉算法算s-d的最短路径，并返回该路径和值
        # print("Start Dijstra Path……")
        path = []  # 用来存储s-d的最短路径
        n = len(network)  # 邻接矩阵维度，即节点个数
        fmax = float('inf')
        w = [[0 for _ in range(n)] for j in range(n)]  # 邻接矩阵转化成维度矩阵，即0→max
        book = [0 for _ in range(n)]  # 是否已经是最小的标记列表
        dis = [fmax for i in range(n)]  # s到其他节点的最小距离
        book[s - 1] = 1  # 节点编号从1开始，列表序号从0开始   ##从s点开始便历
        midpath = [-1 for i in range(n)]  # 上一跳列表
        for i in range(n):
            for j in range(n):
                if network[i][j] != 0:
                    w[i][j] = network[i][j]  # 0→max
                else:
                    w[i][j] = fmax
                if i == s - 1 and network[i][j] != 0:  # 直连的节点最小距离就是network[i][j]
                    dis[j] = network[i][j]
        for i in range(n - 1):  # n-1次遍历，除了s节点
            min = fmax
            u = 0
            for j in range(n):
                if book[j] == 0 and dis[j] < min:  # 如果未遍历且距离最小
                    min = dis[j]
                    u = j
            book[u] = 1
            for v in range(n):  # u直连的节 点遍历一遍
                if v == s - 1:
                    dis[v] = 0
                elif dis[v] > dis[u] + w[u][v]:
                    dis[v] = dis[u] + w[u][v]
                    midpath[v] = u + 1  # 上一跳更新
        j = d - 1  # j是序号
        path.append(d)  # 因为存储的是上一跳，所以先加入目的节点d，最后倒置
        while (midpath[j] != -1):
            path.append(midpath[j])
            j = midpath[j] - 1
        path.append(s)
        path.reverse()  # 倒置列表
        # print("path:", path)
        # print(midpath)
        # print("dis:", dis)
        a = dis[d - 1]  # 下标要减一
        # print('ij:', a)
        # return a,path

        return path


         # 迪杰斯特拉算法算s-d的最短路径，并返回该路径和值

    def init_Population(self):  # 反向策略初始化种群
        ###群数 长度  维度
        self.fitness1=[]
        self.fitness2=[]
        self.pop_l1 = []
        self.pop_t1 = []
        self.pbest1 = []
        self.population = np.zeros((self.pop, self.length, self.dim), dtype=np.int)
        self.V = np.zeros((self.pop, self.length, self.dim), dtype=np.int)
        self.population1 = np.zeros((2*self.pop, self.length,self.dim),dtype= np.int)
        self.V1=np.zeros((2*self.pop, self.length,self.dim),dtype= np.int)
        self.choice1 = [i for i in range(self.dim + 1)]
        # print("self.choice1:",self.choice1)
        # self.choice1=[0,1,2,3,4,5]   ###二维基因记得添加0 ##对应index
        choice2=[-1,0,1]
        for i in range(self.pop):
            # for j in range(self.length):
            j=0
            while j<self.length:
                for k in range(self.dim):
                    self.population1[i][j][k] = random.choice(self.choice1)
                    self.V1[i][j][k] = random.choice(choice2)# 初始化各个粒子的速度
                if np.all(self.population1[i][j]==0)==True:
                    continue
                else:
                    j+=1
        # print(self.population)

        for i in range(self.pop,2*self.pop):
            for j in range(self.length):
                for k in range(self.dim):
                    # self.population1[i][j][k]=(random.random()*self.dim+self.population1[i-self.pop][j][k])
                    self.population1[i][j][k] =self.dim-self.population1[i - self.pop][j][k]
                    self.V1[i][j][k]=self.V1[i-self.pop][j][k]
                    if self.population1[i][j][k] > self.choice1[-1]:
                        self.population1[i][j][k] = self.population1[i][j][k] - self.choice1[-1]
                    elif self.population1[i][j][k] < self.choice1[0]:
                        self.population1[i][j][k]=-self.population1[i][j][k]
        # print(self.population1)

        for i in range(2*self.pop):
            position1 = self.population1[i]
            a1 = self.com_parallel(position1)
            self.pop_l1.append(a1[0])  # [i][0] = f1
            self.pop_t1.append(a1[1])
        tmax1, tmin1 = self.max_min(self.pop_t1)
        self.tmax1 = tmax1
        print(self.tmax1)
        lmax1, lmin1 = self.max_min(self.pop_l1)
        self.lmax1 = lmax1
        print(self.lmax1)
        for i in range(2*self.pop):
            """处理一下gene错误的问题，避开float(inf)"""
            if self.pop_l1[i] == float('inf') or self.pop_t1[i] == float('inf'):
                self.fitness1.append(float('inf'))
            else:
                gfit1 = self.alpha * (self.pop_l1[i] / self.lmax1) + (1 - self.alpha) * (self.pop_t1[i] / self.tmax1)
                # fit=self.alpha*(self.pop_l[i]/self.lmax)+(1-self.alpha)*(self.pop_t[i]/self.tmax)
                self.fitness1.append(gfit1)
                ##第一次的个体历史最优及基因
                self.pbest1.append((gfit1, self.population1[i]))
        a1 = np.array(self.fitness1)
        b1 = np.argsort(a1)  ###从小到大的坐标序号
        for i in range(self.pop):
            self.population[i]=self.population1[b1[i]]
            self.V[i] = self.V1[b1[i]]

       # print(self.pbest1)




    def init_task(self):
        self.ing_task=[]
        for i in range(len(self.work_size_total[self.ing_group])):
            task=[]
            for j in range(len(self.work_size_total[self.ing_group][i])):
                task.append(self.work_size_total[self.ing_group][i][j])
            self.ing_task.append(task)  ##任务数量大小确定

    def max_min(self,list):
        min = list[0]
        max = list[0]
        for x in list:
            """
            处理一下gene错误的问题，避开float(inf)
            """
            # print(x)
            if x==float('inf'):
                continue
            else:
                if x < min:
                    min = x
                if x > max:
                    max = x
        return max, min
    # a, b = max_min(fitness)
    # print(a, b)

    def com_fit(self):
        self.fitness = []
        self.pop_l = []
        self.pop_t = []
        self.pbest = []
        self.gbest_fitness = []
        self.iter_bfit = []
        self.gbest = []
        self.g_fitness_index = []
        for i in range(self.pop):
            position = self.population[i]
            a = self.com_parallel(position)
            self.pop_l.append(a[0])  # [i][0] = f1
            self.pop_t.append(a[1])
        tmax, tmin = self.max_min(self.pop_t)
        self.tmax = tmax
        print(self.tmax)
        lmax, lmin = self.max_min(self.pop_l)
        self.lmax = lmax
        print(self.lmax)
        for i in range(self.pop):
            """处理一下gene错误的问题，避开float(inf)"""
            if self.pop_l[i] == float('inf') or self.pop_t[i] == float('inf'):
                self.fitness.append(float('inf'))
            else:
                gfit = self.alpha * (self.pop_l[i] / self.lmax) + (1 - self.alpha) * (self.pop_t[i] / self.tmax)
                # fit=self.alpha*(self.pop_l[i]/self.lmax)+(1-self.alpha)*(self.pop_t[i]/self.tmax)
                self.fitness.append(gfit)
                ##第一次的个体历史最优及基因
                self.pbest.append((gfit, self.population[i]))
        # 得到各个个体的适应度值
        a = np.array(self.fitness)
        b = np.argsort(a)  ###从小到大的坐标序号
        self.gbest_fitness = self.fitness[b[0]]  # 种群中最佳适应度
        self.iter_bfit.append(self.fitness[b[0]])  ###len:迭代次数
        self.g_fitness_index = b[0]  ##种群中最佳适应度在坐标中序号
        self.gbest = self.population[b[0]]  # 最佳适应度在整个整体的最优位置

    def com_fit1(self):
        self.fitness = []
        self.pop_l = []
        self.pop_t = []
        for i in range(self.pop):
            position = self.population[i]
            a = self.com_parallel(position)
            self.pop_l.append(a[0])  # [i][0] = f1
            self.pop_t.append(a[1])
        for i in range(self.pop):
            """处理一下gene错误的问题，避开float(inf)"""
            if self.pop_l[i] == float('inf') or self.pop_t[i] == float('inf'):
                self.fitness.append(float('inf'))
            else:
                gfit = self.alpha * (self.pop_l[i] / self.lmax) + (1 - self.alpha) * (self.pop_t[i] / self.tmax)
                # fit=self.alpha*(self.pop_l[i]/self.lmax)+(1-self.alpha)*(self.pop_t[i]/self.tmax)
                self.fitness.append(gfit)
                ##第一次的个体历史最优及基因
                self.pbest.append((gfit, self.population[i]))
            # 得到各个个体的适应度值
            #####更新粒子个体历史最优
            if self.pbest[i][0] > gfit:
                self.pbest[i] = (gfit, self.population[i])
        a = np.array(self.fitness)
        b = np.argsort(a)  ###从小到大的坐标序号
        ##当前种群最优
        min_fitness = self.fitness[b[0]]
        self.iter_bfit.append(min_fitness)  ###len:迭代次数
        ##更新种群历史最优
        if min_fitness < self.gbest_fitness:
            self.gbest_fitness = min_fitness
            self.gbest = self.population[b[0]]  # 初始化整个整体的最优粒子


    def com_parallel(self,gene):
        if_break=False
        self.init_task()###只有数据大小
        self.init_task_infor=[]
        self.init_task_distance_infor=[]#保存链路距离大小
        ###（数据大小，路径）没有初始AP
        for i in range(self.bs_num):
            ###任务分份,处理
            AP_corresponding=gene[i]
            task_part = 0
            nonull_index=[]
            for z in range(self.dim):
                if AP_corresponding[z]!=0:
                    task_part+=1
                    nonull_index.append(z)
            if task_part==0:
                print('gene error!')
                if_break=True
                break
            part_numi=int(len(self.ing_task[i])/task_part)
            #[20,50,70]任务分份后对应的每一份的任务索引区间
            task_up=0
            part_num=[]
            for r in range(task_part):
                if r !=task_part-1:
                    part_num.append(task_up+part_numi)
                    task_up += part_numi
                else:
                    part_num.append(len(self.ing_task[i]))

            ###去掉为0的基因位
            new_task_dist = []
            for e in range(len(nonull_index)):
                new_task_dist.append(AP_corresponding[nonull_index[e]])
            ###########任务排序
            s=np.array(self.ing_task[i])
            q=s.argsort()
            ####3
            task_infor=[]
            distance_infor=[]
            a=0
            dis=0
            for j in range(len(self.ing_task[i])):
                if j<part_num[a]:
                    path = self.Dijkstra(self.matrix, i + 1, self.edge_index[nonull_index[a] - 1])  ##注意序号
                    #path=[20]
                    distance=0
                    if (len(path))>1:
                       for m in range(len(path)-1):
                           distance+=(np.sqrt(((self.bs_local[path[m+1]-1][0])-(self.bs_local[path[m]-1][0]))**2+((self.bs_local[path[m+1]-1][1])-(self.bs_local[path[m]-1][1]))**2))
                    taskj_infor=(self.ing_task[i][j],path[1:]) #(1, [2, 32, 5, 4, 5])
                    distancej_infor=(self.ing_task[i][j],path,distance)
                    distance_infor.append(distancej_infor)
                    task_infor.append(taskj_infor)
                    if j+1==part_num[a]:
                        a+=1
            self.init_task_infor.append(task_infor)
            self.init_task_distance_infor.append(distance_infor)
        # print('len(self.init_task_infor):',len(self.init_task_infor))
        # print(len(self.init_task_infor[0]))
        # print(len(self.init_task_infor[0][0]))
        ####
        self.edge_task = [[] for _ in range(self.edge_num)]
        total_ap_tran=0
        remnant=[0 for _ in range(self.bs_num)]  ##上个时隙处理的任务
        null_num=0
        #数组模拟队列
        while null_num < self.bs_num:
            if if_break:
                break

            all_ap_full=[False for _ in range(self.bs_num)]  ##判断时间是否为时隙
            all_ap_end_task = [0 for _ in range(self.bs_num)]  ##判断最后所有时间的长度

            for i in range(self.bs_num):
                # print('AP(i+1):',i+1)
                #达到服务器
                if self.init_task_infor[i]!=[]:
                    #### print('len(self.init_task_infor[i]):',len(self.init_task_infor[i]))
                    # for a in range(len(self.init_task_infor[i])):###剪掉路径为空的任务
                    #     ##判断gene[i]在edge_task中序号
                    #
                    #     if self.init_task_infor[i][0][1]==[]:  #路径中为空
                    #         self.edge_task[index].append(self.init_task_infor[i][0][0])
                    #             #更新原AP中的任务信息集
                    #         self.init_task_infor[i].remove(self.init_task_infor[i][0])
                    #下一跳AP
                    sum_task_ing = remnant[i]
                    for j in range(len(self.init_task_infor[i])):  ###
                        k = 0
                        if remnant[i]!=0 and j==0:
                            sum_task_ing +=self.init_task_infor[i][j][0]-remnant[i]
                        else:
                            sum_task_ing+=self.init_task_infor[i][j][0]
                        # print('sum_task_ing',sum_task_ing)
                        if sum_task_ing<self.tran_v:
                            if j<len(self.init_task_infor[i])-1:
                                if sum_task_ing + self.init_task_infor[i][j + 1][0] > self.tran_v:
                                    save=self.tran_v
                                    is_nofull=False
                                    k = j
                                    break
                                elif sum_task_ing+self.init_task_infor[i][j+1][0]<=self.tran_v:
                                    if j+1==len(self.init_task_infor[i])-1:
                                        is_nofull = True
                                        k = j + 1
                                        save=sum_task_ing+self.init_task_infor[i][j+1][0]
                                        break
                                    else:
                                        continue
                            else:  ##最后一位
                                k = j
                                save= sum_task_ing
                                is_nofull = True
                                break
                        else:
                            is_nofull=True   ###根据速率计算
                            save=self.tran_v
                    # print('k:',k)
                    for b in range(k+1):  #K为索引
                        # print(b)
                        # print(self.init_task_infor[i])
                        #### 去掉任务信息中下一跳目标AP
                        aim_index = self.init_task_infor[i][0][1][0]  # 下一跳目的
                        task_infor = self.init_task_infor[i][0]
                        task_next_index =task_infor[1][0]
                        task_infor = (task_infor[0], task_infor[1][1:])
                        ###更新原AP中任务信息
                        self.init_task_infor[i].remove(self.init_task_infor[i][0])
                        ##判断
                        if task_infor[1]!=[]:
                            self.init_task_infor[aim_index-1].append(task_infor) ##目标AP更新任务信息
                        else:
                            for p in range(len(self.edge_index)):
                                if task_next_index == self.edge_index[p]:
                                    next_index= p
                                    break
                            ###AP的序号和索引号差1
                            self.edge_task[next_index].append(task_infor[0])
                    ##1个时隙没有处理完的任务
                    remnant[i]=self.tran_v-save
                all_ap_full[i]=is_nofull
                all_ap_end_task[i]=save
            # print('all_ap_full:',all_ap_full)
            # print('all_ap_full_task:', all_ap_end_task)
            if False in all_ap_full:
                ##全部AP的最长转发时延
                total_ap_tran+=1
            else:
                short_t=[]
                for i in range(self.bs_num):
                    short_t.append(all_ap_end_task[i]/self.tran_v)
                total_ap_tran +=max(short_t)
            ##
            null_num=0
            for i in range(self.bs_num):   ##判断所有AP中是否还有任务
                # print('self.init_task_infor[i]:', self.init_task_infor[i])
                if self.init_task_infor[i] == []:
                    null_num+=1
                    continue
            # print(null_num)
        ###
        if if_break:
            f1=float('inf')
            f2=float('inf')
        else:
            f2=total_ap_tran
            #############????????????????????
            resource = 0
            for i in range(self.edge_num):
                # print('sum(self.edge_task[i]):', sum(self.edge_task[i]))
                resource += sum(self.edge_task[i])
            resource_ave = resource / self.edge_num
            # print('resource_ave:', resource_ave)
            resource = []
            # for i in range(self.edge_num):
            #     if ((sum(self.edge_task[i])*self.s*1024*1024*8*20)>
            self.exec = [[] for _ in range(self.edge_num)]
            max1=0
            min1=0
            for i in range(self.edge_num):  # 计算运行时延
                for j in range(len(self.edge_task[i])):
                    t_server = (self.edge_task[i][j] * 1024 * 8) / ((self.s1) / self.s)
                    if  t_server>max1:
                        max1=t_server
                    elif t_server<min1:
                        min1=t_server
                    self.exec[i].append(t_server)
            for i in range(self.edge_num):
                resource_i = (sum(self.edge_task[i]) - resource_ave) ** 2
                resource.append(resource_i)
            L_balance = math.sqrt((sum(resource)) / self.edge_num)
            f1 =L_balance
            f2=f2+(max1+min1)/2
            print('f1,f2:', f1, f2)
            t_server = 0
           #self.edge_task = [[] for _ in range(self.edge_num)]
            # self.exec = [[] for _ in range(self.edge_num)]
            # for i in range(self.edge_num):#计算运行时延
            #     for j in range(len(self.edge_task[i])):
            #         t_server = (self.edge_task[i][j]*1024*1024*8)/ ((self.mec/500) / self.s)
            #         self.exec[i].append(t_server)

        return [f1,f2]

    # 更新速度，根据公式V(t+1)=w*V(t)+c1*r1*(pbest_i-xi)+c1*r1*(gbest_xi)
    def velocity_gene_update(self,V, X, pbest, gbest):###一个粒子的速度更新
        w = ((self.w_ini - self.w_end) * (((self.max_iter - self.now_iter + 1)**2) / (self.max_iter)**2)) + self.w_end
        c1=self.cmin+(self.cmax-self.cmin)*np.cos((self.now_iter/self.max_iter)*((math.pi)/2))
        c2 = self.cmin + (self.cmax - self.cmin) * np.sin((self.now_iter / self.max_iter) * ((math.pi) / 2))
        r1 = np.random.random((self.length, self.dim))  # 该函数表示成size行 1列的浮点数，浮点数都是从0-1中随机。
        r2 = np.random.random((self.length, self.dim))

        for i in range(self.length):
            j=0
            is_all_0=0
            while j <self.dim:
            # for j in range(self.dim):
            #     gbest[i][j]=gbest[i][j]+np.random.standard_cauchy() * 0.1

                vi=w*V[i][j]+c1 * r1[i][j] *(pbest[i][j]-X[i][j])+c2 * r2[i][j] * (gbest[i][j] - X[i][j])
                beta = 1.5
                sigma = (gamma(1 + beta) * np.sin(np.pi * beta / 2) / (
                    gamma((1 + beta) / 2) * beta * 2 ** ((beta - 1) / 2))) ** (1 / beta)
                u = np.random.normal(0, sigma, self.dim)
                v = np.random.normal(0, 1, self.dim)
                step = u / np.abs(v) ** (1 / beta)
                vi = vi + 0.01 * (pbest[i][j] - X[i][j]) * step[j]
                vi=int(vi)
                # 这里是一个防止速度过大的处理，怕错过最理想值
                if vi < -self.max_val:
                    vi= -self.max_val
                elif vi > self.max_val:
                    vi= self.max_val
                xi = X[i][j] + vi
                if xi<self.choice1[0]:
                    xi=-xi
                elif xi>self.choice1[-1]:
                    xi = xi-self.choice1[-1]
                if xi==0:
                    is_all_0+=1
                if xi in self.choice1:
                    X[i][j]=xi
                    j+= 1
                    continue
                else:
                    r1[i][j] = random.random()
                    r2[i][j] = random.random()
                if is_all_0==self.dim:
                    j=0
                    is_all_0 = 0
                    r1 = np.random.random((self.length, self.dim))  # 该函数表示成size行 1列的浮点数，浮点数都是从0-1中随机。
                    r2 = np.random.random((self.length, self.dim))
        # V = w * V + c1 * r1 * (pbest - X) + c2 * r2 * (gbest - X)  # 注意这里得到的是一个矩阵
        return V,X

    def update(self):
        self.new_popu = np.zeros((self.pop, self.length,self.dim), dtype=np.int)
        self.new_V=np.zeros((self.pop, self.length,self.dim), dtype=np.int)
        a = np.array(self.fitness)
        ind = np.argsort(a)  ###从小到大的坐标序号
        # 添加前5%
        for i in range(int((1 / 20) * self.pop)):
            self.new_popu[i] = self.population[ind[i]]
            self.new_V[i]=self.V[ind[i]]
        #将前95%进行更新
        for i in range(int((19/ 20) * self.pop)):
            X=self.population[ind[i]]
            V=self.V[ind[i]]
            pbest=self.pbest[ind[i]][1]
            gbest=self.gbest
            V,X=self.velocity_gene_update(V,X,pbest, gbest)
            # X=self.gene_updata(X, V)
            self.new_popu[i+int((1 /20) * self.pop)]=X
            self.new_V[i+int((1 / 20) * self.pop)]=V
        ###
        self.V=np.zeros((self.pop, self.length,self.dim), dtype=np.int)
        self.population = np.zeros((self.pop, self.length,self.dim), dtype=np.int)
        for i in range(self.pop):
            self.population[i] = self.new_popu[i]
            self.V[i]= self.new_V[i]

    def save(self):
        #####
        a = np.array(self.fitness)
        b = np.argsort(a)  ###从小到大的坐标序号
        self.total_L.append(self.pop_l[b[0]])
        self.total_t.append(self.pop_t[b[0]])
        print(' self.iter_bfit:', self.iter_bfit)
        with open('array1.txt', 'w') as f:
            for item in self.iter_bfit:
                f.write("%s\n" % item)
        print('self.total_L:', self.total_L)
        print('self.total_t:', self.total_t)


    def save_excel1(self):
        ##save data to excel
        wb1 = Workbook()
        # 激活sheet
        sheet1 = wb1.active
        sheet1.title = 'GAPSO'
        for a in range(len(self.total_L)):
            sheet1.cell(a + 1, 1).value = self.total_L[a]  # x
            sheet1.cell(a + 1, 2).value = self.total_t[a]  # y
        wb1.save('GAPSOresult0505.xlsx')

    def run(self):
        self.read()
        for j in range(self.several_groups):
            self.init_Population()  # 初始化种群，选择交叉变异，生成子代种群
            self.iter_bfit = []
            """单独计算某一个  总任务数量  的结果 """
            if j==1:
                for k in range(self.max_iter):
                    self.now_iter = k
                    print('finished :,k:', self.ing_group, k)
                    self.init_task()
                    if k == 0:
                        self.com_fit()
                        # print(self.lmax,self.tmax)
                    else:
                        self.com_fit1()
                        # print(self.lmax, self.tmax)
                    #self.crossover()
                    # self.mutation()
                    self.update()
                self.save()
            print('finished :', self.ing_group)
            self.ing_group += 1
        #self.save_excel1()


if __name__ == '__main__':
    """AP数量，mes数量，种群数量，迭代次数"""
    A=GA_PSO(18,5,120,100)  ##length, dim,pop, max_iter  pop必须偶数
    A.run()

